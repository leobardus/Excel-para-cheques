import pandas as pd
import os
import numpy as np
import json
import sys
import openpyxl
import xlsxwriter
import streamlit as st
from io import BytesIO
from collections import defaultdict
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader

# --- CONFIGURACIÓN DE USUARIOS ---
# HASH generado para la contraseña 'Lajefa25' usando bcrypt
HASH_CONTRASENA_CECI = '$2b$12$SbBRF2XUQmoXREIeLfqbrejn2WMrBuj5Zn7sMlnAL58oJW6O.jw.O' 

config_yaml = {
    'cookie': {
        'expiry_days': 30,
        'key': 'some_signature_key_2024', # CLAVE SECRETA: Cámbiala por una clave única y larga
        'name': 'processor_cookie'
    },
    'credentials': {
        'usernames': {
            'Ceci': { # USUARIO: Ceci
                'email': 'ceci@empresa.com',
                'name': 'Ceci (La Jefa)',
                'password': HASH_CONTRASENA_CECI
            }
        }
    }
}


# --- ⚙️ FUNCIÓN DE CARGA DE CONFIGURACIÓN DE DATOS ---
def cargar_configuracion(nombre_archivo="config.json"):
    """Carga los parámetros de configuración desde un archivo JSON."""
    try:
        with open(nombre_archivo, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except FileNotFoundError:
        st.error(f"Error: El archivo '{nombre_archivo}' no existe.")
        st.stop()
    except (json.JSONDecodeError, ValueError) as e:
        st.error(f"Error de formato en el archivo JSON: {e}")
        st.stop()

# Cargar la configuración global al inicio
CONFIG = cargar_configuracion()
if not CONFIG:
    st.stop()


# --- FUNCIÓN DE AYUDA ---
def hacer_nombres_unicos(columnas):
    """Genera nombres de columna únicos, útil tras la lectura de Excel."""
    visto = defaultdict(int)
    nuevas_columnas = []
    for col in columnas:
        col_str = str(col)
        if col_str in visto:
            visto[col_str] += 1
            nuevas_columnas.append(f"{col_str}.{visto[col_str]}")
        else:
            nuevas_columnas.append(col_str)
            visto[col_str] = 0
    return nuevas_columnas

# -----------------------------------------------
# --- FUNCIÓN DE PROCESAMIENTO CENTRALIZADA ---
# -----------------------------------------------

def procesar_archivo(uploaded_file, config):
    """
    Función principal que realiza el procesamiento de datos, adaptada para Streamlit.
    Devuelve un objeto BytesIO con el Excel procesado.
    """
    
    # --- Parámetros de Configuración ---
    CADENAS_A_ANULAR = config["CADENAS_A_ANULAR"]
    INDICES_A_CONSERVAR = config["INDICES_A_CONSERVAR"]
    INDICE_COLUMNA_ORIGEN = config["INDICE_COLUMNA_ORIGEN"]
    INDICE_COLUMNA_DESTINO = config["INDICE_COLUMNA_DESTINO"]
    INDICE_A_ELIMINAR_DESPUES_DE_COPIA_2 = config["INDICE_A_ELIMINAR_DESPUES_DE_COPIA_2"]
    MAPEO_RENOMBRE = config["MAPEO_RENOMBRE"]
    
    input_file_bytes = BytesIO(uploaded_file.getvalue())
    output = BytesIO()

    try:
        # 1. --- Descombinar Celdas con openpyxl (en memoria) ---
        libro = openpyxl.load_workbook(input_file_bytes)

        for nombre_hoja in libro.sheetnames:
            hoja = libro[nombre_hoja]
            rangos_a_descombinar = list(hoja.merged_cells.ranges)
            for rango_obj in rangos_a_descombinar:
                hoja.unmerge_cells(str(rango_obj))
                
        temp_file = BytesIO()
        libro.save(temp_file)
        temp_file.seek(0)
        
        # 2. --- Estructuración de Columnas y Filas con Pandas ---
        datos_limpios = {}
        excel_data = pd.read_excel(temp_file, sheet_name=None, header=None, engine='openpyxl')
        
        for nombre_hoja, df in excel_data.items():
            
            # 2a. Eliminar las primeras 10 filas
            if df.shape[0] > 10:
                df_procesado = df.iloc[10:].reset_index(drop=True)
            else:
                df_procesado = df

            # 2b. Retención de Columnas por Índice
            df_intermedio = df_procesado.iloc[:, INDICES_A_CONSERVAR]
            df_intermedio = df_intermedio.dropna(axis=0, how='all').reset_index(drop=True)

            if not df_intermedio.empty:
                df_final = df_intermedio.copy()
                df_final.columns = df_final.iloc[0]
                df_final = df_final[1:].reset_index(drop=True)
                
                # Normalización de encabezados
                raw_columns = df_final.columns.tolist()
                normalized_columns = [
                    str(col).strip().lower().replace(' ', '_').replace('.', '').replace('-', '_') 
                    for col in raw_columns
                ]
                nombres_unicos = hacer_nombres_unicos(normalized_columns)
                df_final.columns = nombres_unicos
                df_final = df_final.reset_index(drop=True)

                # 3b. ANULAR CONTENIDO BASADO EN PATRONES
                columna_para_filtrar = df_final.columns[0]
                mascara_anular = pd.Series([False] * len(df_final))
                
                for patron in CADENAS_A_ANULAR:
                    mascara_patron = df_final[columna_para_filtrar].astype(str).str.contains(patron, na=False)
                    mascara_anular = mascara_anular | mascara_patron

                df_final.loc[mascara_anular, columna_para_filtrar] = np.nan
            
                # 3c. COPIAR COLUMNA DE ORIGEN A COLUMNA DE DESTINO
                columna_destino_nombre = df_final.columns[INDICE_COLUMNA_DESTINO]
                columna_origen_nombre = df_final.columns[INDICE_COLUMNA_ORIGEN]

                df_final[columna_destino_nombre] = np.where(
                    df_final[columna_origen_nombre].notna(), 
                    df_final[columna_origen_nombre].astype(str), 
                    df_final[columna_destino_nombre]
                )

                # 3d. ELIMINAR COLUMNA
                if INDICE_A_ELIMINAR_DESPUES_DE_COPIA_2 < len(df_final.columns):
                    nombre_columna_a_eliminar_1 = df_final.columns[INDICE_A_ELIMINAR_DESPUES_DE_COPIA_2]
                    df_final = df_final.drop(columns=[nombre_columna_a_eliminar_1])
                
                # 3e. REORDENAR
                columnas = df_final.columns.tolist()
                columna_a_mover = columnas.pop(0)
                columnas.insert(2, columna_a_mover) 
                df_final = df_final[columnas]
                
                # 3f. RENOMBRE DEFINITIVO DE COLUMNAS 
                df_final = df_final.rename(columns=MAPEO_RENOMBRE)
                
                # 3g. RELLENO HACIA ABAJO (FFILL) Y ADICIÓN DEL PREFIJO '0'
                if "Codigo" in df_final.columns:
                    df_final["Codigo"] = df_final["Codigo"].ffill()
                    df_final["Codigo"] = df_final["Codigo"].astype(str).str.replace(r'\.0$', '', regex=True)
                    df_final["Codigo"] = df_final["Codigo"].replace('nan', np.nan) 
                    df_final["Codigo"] = np.where(df_final["Codigo"].notna(), 
                                                  "0" + df_final["Codigo"], 
                                                  df_final["Codigo"])
                
                # 3h. LIMPIEZA ADICIONAL y eliminación de filas vacías
                df_final = df_final.replace(r'^\s*$', np.nan, regex=True)
                df_final = df_final.dropna(axis=0, thresh=2).reset_index(drop=True)
                
                # 3j. CONVERSIÓN DE TIPOS PARA FORMATO (Eliminar hora)
                if "Fecha Vto" in df_final.columns:
                    df_final["Fecha Vto"] = pd.to_datetime(df_final["Fecha Vto"], errors='coerce')
                    df_final["Fecha Vto"] = df_final["Fecha Vto"].dt.date 
                    
                if "Fecha Em" in df_final.columns:
                    df_final["Fecha Em"] = pd.to_datetime(df_final["Fecha Em"], errors='coerce')
                    df_final["Fecha Em"] = df_final["Fecha Em"].dt.date
                    
                for col in ["Importe", "Mora"]:
                    if col in df_final.columns:
                        df_final[col] = pd.to_numeric(df_final[col], errors='coerce')
                
                # 3k. REORDENAMIENTO FINAL
                columnas_finales = df_final.columns.tolist()
                nuevo_orden = []
                
                orden_deseado = ['Codigo', 'Proveedor', 'Fecha Em', 'Fecha Vto', 'Importe', 'Mora']
                
                for nombre in orden_deseado:
                    if nombre in columnas_finales:
                        nuevo_orden.append(nombre)
                
                for col in columnas_finales:
                    if col not in nuevo_orden:
                        nuevo_orden.append(col)
                
                df_final = df_final[nuevo_orden]

            else:
                df_final = df_intermedio

            datos_limpios[nombre_hoja] = df_final
            
        # 4. --- Guardar el Archivo Final con Formato (a BytesIO) ---

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            for sheet_name, df_sheet in datos_limpios.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Definición de formatos
                date_format_corta = workbook.add_format({'num_format': 'dd/mm/yyyy'}) 
                currency_format = workbook.add_format({'num_format': '$ #,##0.00'})
                integer_format = workbook.add_format({'num_format': '#,##0'}) 

                final_columns = df_sheet.columns.tolist()

                # Aplicar formato de fecha, moneda e entero
                for col_name, fmt in [
                    ("Fecha Em", date_format_corta), 
                    ("Fecha Vto", date_format_corta), 
                    ("Importe", currency_format), 
                    ("Mora", integer_format)
                ]:
                    if col_name in final_columns:
                        col_index = final_columns.index(col_name)
                        worksheet.set_column(col_index, col_index, 12, fmt)

        output.seek(0)
        return output

    except Exception as e:
        st.error(f"Ocurrió un error inesperado durante el procesamiento: {e}")
        return None

# -------------------------------------------------------------------
# --- INTERFAZ DE STREAMLIT CON AUTENTICACIÓN ---
# -------------------------------------------------------------------

def app_content():
    """Contenido de la aplicación visible solo después del inicio de sesión."""
    
    st.title("⚙️ Procesador Web de Reportes de Proveedores")
    st.markdown(f"Bienvenido(a), **{st.session_state['name']}**. Sube tu archivo y procesa el reporte.")
    
    # 🚨 NUEVO MENSAJE DE ADVERTENCIA 🚨
    st.warning(
        "⚠️ **¡ATENCIÓN!** Antes de subir el archivo, **copie la información del excel generado por Bejerman en un nuevo archivo de Excel** y utilice ese nuevo archivo para subirlo a la página web."
    )

    # 1. Selector de Archivo
    uploaded_file = st.file_uploader(
        "Sube aquí el archivo .xlsx (Excel)", 
        type=["xlsx"],
        help="Solo archivos de Excel con extensión .xlsx"
    )

    if uploaded_file is not None:
        
        # 2. Botón de Procesamiento
        if st.button("🚀 Procesar Archivo y Descargar"):
            
            with st.spinner('Procesando datos y aplicando formatos... Esto puede tardar unos segundos.'):
                
                nombre_base = os.path.splitext(uploaded_file.name)[0]
                nombre_salida = f"{nombre_base}_PROCESADO_FINAL.xlsx"
                
                # Llamar a la función de procesamiento
                processed_excel = procesar_archivo(uploaded_file, CONFIG)
                
                if processed_excel:
                    st.success("✅ ¡Procesamiento completado con éxito!")
                    
                    # 3. Botón de Descarga
                    st.download_button(
                        label="⬇️ Descargar Archivo Procesado",
                        data=processed_excel,
                        file_name=nombre_salida,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Opcional: Mostrar las primeras filas (solo si es una sola hoja para simplificar)
                    try:
                        processed_excel.seek(0)
                        df_preview = pd.read_excel(processed_excel, sheet_name=0)
                        st.subheader("Vista previa de los primeros 5 registros:")
                        st.dataframe(df_preview.head(), use_container_width=True)
                    except Exception as e:
                        st.warning(f"No se pudo mostrar la vista previa: {e}")


def main():
    st.set_page_config(page_title="Procesador Web de Reportes de Proveedores", layout="centered")

    # Inicializar el autenticador
    authenticator = stauth.Authenticate(
        config_yaml['credentials'],
        config_yaml['cookie']['name'],
        config_yaml['cookie']['key'],
        config_yaml['cookie']['expiry_days']
    )

    # --- Mostrar el formulario de inicio de sesión ---
    name, authentication_status, username = authenticator.login('Inicio de Sesión', 'main')

    if authentication_status:
        # 1. ESTADO: Autenticado
        st.session_state['name'] = name # Guardar el nombre del usuario en la sesión
        st.sidebar.markdown(f"**Bienvenido/a:** {name}")
        authenticator.logout('Cerrar Sesión', 'sidebar') # Botón de cierre en la barra lateral
        app_content() # Mostrar el contenido principal de la aplicación
        
    elif authentication_status is False:
        # 2. ESTADO: Falla de autenticación
        st.error('Nombre de usuario/contraseña incorrectos.')
    
    elif authentication_status is None:
        # 3. ESTADO: No ha intentado o está pendiente
        st.warning('Por favor, ingresa tu nombre de usuario y contraseña para acceder.')


if __name__ == '__main__':
    main()